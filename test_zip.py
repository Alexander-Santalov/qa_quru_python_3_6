import glob
import os
from os.path import basename
from zipfile import ZipFile
import csv
import pytest
from PyPDF2 import PdfReader
from io import TextIOWrapper
from openpyxl import load_workbook


@pytest.fixture()
def clear_dir():
    precondition_directory()


path = os.path.join(os.path.dirname(os.path.abspath(__file__)), 'download')
path_destination = os.path.join(os.path.dirname(os.path.abspath(__file__)), 'resources')
path_zip = os.path.join(path_destination, "asantalov.zip")


def test_create_archive(clear_dir):
    file_dir = os.listdir(path)
    with ZipFile(path_zip, "w") as myzip:
        for file in file_dir:
            add_file = os.path.join(path, file)
            myzip.write(add_file, basename(add_file))
    files = os.listdir(path_destination)
    assert len(files) == 1, f"Неверное количество скаченных файлов {len(files)} не ровно {1}"
    assert "asantalov.zip" == files[0], f"Архив {files[0]} создался с неправильным именем"


def test_read_and_content_csv():
    zf = ZipFile(path_zip)
    with zf.open("username.csv") as csvfile:
        csvfile = csv.reader(TextIOWrapper(csvfile))
        list_csv = []
        for r in csvfile:
            text = "".join(r).replace(";", " ", 3)
            list_csv.append(text)
    assert "johnson81 4081 Craig Johnson" in list_csv, f"В файле отсутсвует информация " \
                                                       f"о пользователе {'johnson81 4081 Craig Johnson'}"
    zf.close()


def test_read_and_content_pdf():
    with ZipFile(path_zip) as zf:
        pdf_file = zf.extract("journal.pdf")
        reader = PdfReader(pdf_file)
        try:
            page = reader.pages[0]
            text = page.extract_text()
            result_filters = text.split("№")
            del result_filters[0]
            for text in result_filters:
                assert "Зона: (2) Auto_test_zone" in text, \
                    f"В результате есть зона отличная от 'Зона: (2) Auto_test_zone'"
        finally:
            os.remove(os.path.join(os.path.dirname(os.path.abspath(__file__)), 'journal.pdf'))
            zf.close()


def test_read_and_content_xlsx():
    zf = ZipFile(path_zip)
    with zf.open("file_example.xlsx") as xlsxfile:
        xlsxfile = load_workbook(xlsxfile)
        sheet = xlsxfile.active
        assert sheet.cell(row=4, column=5).value == "France", f"Ожидаемый результат {'France'}, " \
                                                              f"Фактический {sheet.cell(row=4, column=5).value}"
    zf.close()


def precondition_directory():
    path_file = os.path.join(path_destination, '*.*')
    for file in glob.glob(path_file):
        os.remove(file)
